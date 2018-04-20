#
#
# Title:        translation_strings_tool.py
# Updated:      20-04-2018
# Version:      v2.0.1
# Author:       bRiggin
#
#

#
# imports
#

import openpyxl
import logging
import argparse
import sys
import os
import traceback
import math
import xml.etree.ElementTree as elementTree
import xml.dom.minidom as minidom
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from enum import Enum


class CellType(Enum):
    """ CellType

    Enumeration class that allows tool to utilise a switch statement to determine the it's behaviour (dictated by
    user's selection).
    """
    cell_type = 0
    modifier = 1
    key = 2
    string = 3


DESTINATION_STRING_NOT_DEFINED = "!mp@$$!&L£|P@+h"
XML_TITLE = "strings.xml"
WORKSHEET_TITLE = "Deconstructed Strings"

#
# logger setup
#
logger = logging.getLogger("Android Translation Strings Tool")
logger.setLevel(logging.DEBUG)

handler = logging.StreamHandler()
handler.setLevel(logging.DEBUG)

formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

logger.addHandler(handler)

#
# args
#
parser = argparse.ArgumentParser(prog="Android Translation Strings Tool",
                                 formatter_class=argparse.RawDescriptionHelpFormatter,
                                 description="This tool has been designed to simplify the file "
                                             "deconstruction/construction required\nduring the process of extending an "
                                             "Android project’s User Interface (UI) strings.xml\nfile to different "
                                             "languages.\n\nThe tool interrogates the provided strings.xml file and "
                                             "uses its contents to populated\na structured spreadsheet. The translated "
                                             "UI strings can then can copied into this\nspreadsheet and the tool will "
                                             "reverse the process and create the appropriate\nstrings.xml file for all "
                                             "supplied languages.\n\nWARNING - This tool will automatically overwrite "
                                             "files.")
group = parser.add_mutually_exclusive_group()

group.add_argument("-d", "--deconstruct", action="store_true", help="Deconstructs strings.xml into spreadsheet.")
group.add_argument("-c", "--construct", action="store_true",
                   help="Constructs all required strings.xml files from parsed Excel spreadsheet.")

parser.add_argument("excel_file_name", type=str, help="Excel file name that will be created or is being read from.")
parser.add_argument("source_path", type=str, help="Directory of data source (strings.xml file or spreadsheet).")
parser.add_argument("destination_path", nargs='?', default=DESTINATION_STRING_NOT_DEFINED, type=str,
                    help="Optional, if included, output file(s) will be stored in this directory.")


def main(args):
    """ Main

    Main function of script.

    :param args: Arguments passed by caller.
    """
    try:
        logger.info("Selected mode: " + str(sys.argv[1]))

        if args.deconstruct:
            launch_xml_deconstruction(args.source_path, args.destination_path, args.excel_file_name)

        elif args.construct:
            launch_xml_construction(args.source_path, args.destination_path, args.excel_file_name)
        else:
            logger.warning("Do not recognise mode argument")

    except Exception as exception:
        logger.error(repr(exception) + '\n' + str(exception.args) + '\n' + traceback.format_exc())


def launch_xml_deconstruction(source_path, destin_path, filename):
    """ Launch strings.xml file deconstruction

    Called from main and initiates the deconstruction of the supplied strings.xml file.

    :param source_path:    User provided path of strings.xml file.
    :param destin_path:    Destination path of constructed spreadsheet.
    :param filename:       Filename of created Excel file.
    """
    xml_items = read_xml_file(source_path)
    workbook = openpyxl.Workbook()

    for i in workbook.worksheets:
        workbook.remove(i)

    workbook.create_sheet(WORKSHEET_TITLE)
    worksheet = workbook[WORKSHEET_TITLE]

    populate_worksheet(xml_items, worksheet)

    style_worksheet(worksheet)

    file_extension_included = False

    if '.xlsx' in filename:
        file_extension_included = True
    elif '.xls' in filename:
        logger.warning("This tool is unable to process \'.xls\' files. The created spreadsheet will be saved with \'.xl"
                       "sx\' file extension.")
        filename = filename[:-4]

    if not file_extension_included:
        filename = "{}.xlsx".format(filename)

    if destin_path == DESTINATION_STRING_NOT_DEFINED:
        file_path = os.path.join(source_path, filename)
    else:
        file_path = os.path.join(destin_path, filename)
    workbook.save(file_path)
    logger.info("Excel file successfully saved at: {}".format(file_path))


def launch_xml_construction(source_path, destin_path, filename):
    """ Launch strings.xml file construction

    Called from main and initiates the construction of the all required strings.xml files.

    :param source_path:    User provided path of Excel file.
    :param destin_path:    Destination path of constructed string.xml files.
    :param common_path:    If true, output file directories should be stored in source_path.
    :param filename:       Filename of read Excel file.
    """
    workbook = read_excel_file(source_path, filename)
    worksheet = get_excel_worksheet(workbook, WORKSHEET_TITLE)

    number_of_rows = worksheet.max_row
    number_of_columns = worksheet.max_column

    logger.info("Excel file loaded. {} XML elements identified in {} languages.".
                format(number_of_rows-1, number_of_columns-3))

    column_limit = create_folders(worksheet, number_of_columns, source_path, destin_path)

    for column in range(3, column_limit+1):
        if destin_path == DESTINATION_STRING_NOT_DEFINED:
            directory = r"{}".format(os.path.join(source_path,
                                                  worksheet["{}1".format(get_column_value(column))].value))
        else:
            directory = r"{}".format(os.path.join(destin_path,
                                                  worksheet["{}1".format(get_column_value(column))].value))

        create_xml_file(worksheet, column, number_of_rows, elementTree.Element('resources'), directory)


def read_xml_file(path):
    """ Read XML File

    Exacts all XML elements from the strings.xml file in at the provided path.

    :param path:        User provided path of strings.xml file.
    :return children:   A list of all xml elements.
    """
    try:
        file_path = os.path.join(path, XML_TITLE)
        xml_content = elementTree.parse(file_path)
        xml_root = xml_content.getroot()
        children = list(xml_root)

        logger.info(XML_TITLE + " file successfully loaded, " + str(len(children)) + " items identified.")

        return children

    except FileNotFoundError:
        logger.error("Was unable to find " + XML_TITLE + " in provided path: " + path)
        exit(1)
    except elementTree.ParseError:
        logger.error("Was unable to read " + XML_TITLE +
                     " file, check that file is not empty and is correctly formatted")
        exit(1)
    except Exception as exception:
        error_string_one = str(repr(exception))
        error_string_two = str(exception.args)
        error_string_three = str(traceback.format_exc())
        logger.debug(error_string_one + "\n" + error_string_two + "\n" + error_string_three)
        logger.error("There was an error during processing the parsed " + XML_TITLE + " file.")
        exit(1)


def read_excel_file(path, filename):
    """ Read Excel File

    Loads the Excel file into a openPyXl workbook.

    :param path:        User provided path of Excel file.
    :param filename:    Filename of Excel file.
    """
    try:
        no_file_extension = False

        if '.xlsx' in filename:
            no_file_extension = True
        elif '.xls' in filename:
            logger.error("This tool is unable to process \'.xls\' files. Please ensure that Excel file has \'.xlsx\' "
                         "extension.")
            exit(1)

        if not no_file_extension:
            filename = "{}.xlsx".format(filename)

        if filename in path:
            logger.error("Excel file read failed. Ensure that supplied path does not contain filename.")
            exit(1)
        else:
            workbook = openpyxl.load_workbook(os.path.join(path, filename))
            return workbook

    except FileNotFoundError:
        logger.error("Was unable to find " + filename + " in provided path: " + path)
        exit(1)

    except Exception as exception:
        error_string_one = str(repr(exception))
        error_string_two = str(exception.args)
        error_string_three = str(traceback.format_exc())
        logger.debug(error_string_one + "\n" + error_string_two + "\n" + error_string_three)
        logger.error("There was an error during processing the parsed " + filename + " file.")
        exit(1)


def get_excel_worksheet(workbook, worksheet_title):
    """ Get Excel Worksheet

    Gets openPyXl worksheet from parsed workbook

    :param workbook:        Parsed openPyXl workbook
    :param worksheet_title: Title of worksheet.
    :return:                openPyXl worksheet.
    """
    try:
        worksheet = workbook[worksheet_title]
        logger.info("\"{}\" sheet successfully loaded from Excel file.".format(worksheet_title))
        return worksheet

    except KeyError:
        logger.error("Was unable to find " + worksheet_title + " in Excel file.")
        exit(1)
    except Exception as exception:
        error_string_one = str(repr(exception))
        error_string_two = str(exception.args)
        error_string_three = str(traceback.format_exc())
        logger.debug(error_string_one + "\n" + error_string_two + "\n" + error_string_three)
        logger.error("There was an error while reading " + worksheet_title + " sheet from Excel file.")
        exit(1)


def create_folders(worksheet, columns, source_path, destin_path):
    """ Create Folders

    Creates folders required to store the string.xml files.

    :param worksheet:       opnePyXl worksheet.
    :param columns:         The total number of columns in worksheet
    :param source_path:     Source directory of Excel file.
    :param destin_path:     Destination directory where where folders should be created (if mode is application).
    :return return_index:   Index of last valid column in worksheet.
    """
    return_index = 0
    for column in range(3, columns):
        try:
            language = worksheet["{}1".format(get_column_value(column))].value
            if destin_path == DESTINATION_STRING_NOT_DEFINED:
                directory = r"{}".format(os.path.join(source_path, language))
            else:
                directory = r"{}".format(os.path.join(destin_path, language))
            if not os.path.exists(directory):
                os.makedirs(directory)
                logger.info("New folder: {}, created at: {}".format(language, directory))
            else:
                logger.info("{} folder already exists at: {}".format(language, directory))
            return_index = column
        except TypeError:
            pass
        except Exception as exception:
            error_string_one = str(repr(exception))
            error_string_two = str(exception.args)
            error_string_three = str(traceback.format_exc())
            logger.debug(error_string_one + "\n" + error_string_two + "\n" + error_string_three)
            logger.error(
                "There was an error while trying to detect/create the required directory to store output files")
            exit(1)
    return return_index


def populate_worksheet(xml_elements, worksheet):
    """" Populate Workbook

    Populates a openPyXl workbook with all XML elements.

    :param xml_elements:   The XML elements that have been taken from the parsed strings.xml.
    :param worksheet:      The openPyXl worksheet that is to be populated with the XML elements.
    :return worksheet:     Same worksheet parsed to function but now containing all XML elements.
    """
    try:
        excel_row_index = 1
        for element in xml_elements:
            excel_row_index += 1
            populate_cell(worksheet, excel_row_index, CellType.cell_type, element.tag)
            child_elements = list(element)
            # XML element has no child elements.
            if len(child_elements) == 0:
                populate_cell(worksheet, excel_row_index, CellType.key, element.attrib["name"])
                populate_cell(worksheet, excel_row_index, CellType.string, element.text)
            else:
                populate_cell(worksheet, excel_row_index, CellType.key, element.attrib["name"])
                # Element is string-array or plurals
                if element.tag == "plurals" or element.tag == "string-array":
                    for item in child_elements:
                        excel_row_index += 1
                        populate_cell(worksheet, excel_row_index, CellType.cell_type, item.tag)
                        if element.tag == "plurals":
                            populate_cell(worksheet, excel_row_index, CellType.key, item.attrib["quantity"])
                        # Element has no string modifiers
                        if len(item) == 0:
                            populate_cell(worksheet, excel_row_index, CellType.string, item.text)
                        else:
                            mod_string, ui_string = derive_modifiers_and_string(item)
                            populate_cell(worksheet, excel_row_index, CellType.modifier, mod_string)
                            populate_cell(worksheet, excel_row_index, CellType.string, ui_string)
                # Element is a string with modifiers (<b></b>, <u></u>, etc).
                else:
                    mod_string, ui_string = derive_modifiers_and_string(element)
                    populate_cell(worksheet, excel_row_index, CellType.modifier, mod_string)
                    populate_cell(worksheet, excel_row_index, CellType.string, ui_string)
        logger.info("All XML elements successfully loaded into Excel worksheet.")

        return worksheet
    except Exception as exception:
        error_string_one = str(repr(exception))
        error_string_two = str(exception.args)
        error_string_three = str(traceback.format_exc())
        logger.debug(error_string_one + "\n" + error_string_two + "\n" + error_string_three)
        logger.error("There was an error during processing the parsed " + XML_TITLE + " file.")
        exit(1)


def populate_cell(worksheet, row, cell_type, value):
    """ Populate Cell

    Function writes the parsed value into the parsed worksheet in a co-ordinate that is determined by the parsed
    type and row.

    :param worksheet:   Parsed openPyXl worksheet that value is to be written in.
    :param row:         The row index of the cell to be written in.
    :param cell_type:   CellType Enum instance that describes the type of cell and therefore it's X co-ordinate.
    :param value:       The value to be written.
    """
    if cell_type == CellType.cell_type:
        co_ordinate = "A{}".format(row)
        worksheet[co_ordinate] = value
    elif cell_type == CellType.modifier:
        co_ordinate = "B{}".format(row)
        worksheet[co_ordinate] = value
    elif cell_type == CellType.key:
        co_ordinate = "C{}".format(row)
        worksheet[co_ordinate] = value
    elif cell_type == CellType.string:
        co_ordinate = "D{}".format(row)
        worksheet[co_ordinate] = value
    else:
        logger.warning("Unrecognised type for XML element: {}".format(row))


def element_has_key(element, key):
    """ Element Has key

    Returns True if "key" is in "element's" attributes, else False.

    :param element:     Xml element that is to be searched.
    :param key:         The key to search for in the element.
    :returns            boolean
    """
    if key in element.attrib:
        return True
    else:
        return False


def is_deepest_item(element):
    """ Is Deepest Item

    Returns True if parsed XML element as no child elements, else returns False.

    :param element:     Parsed XML element.
    :returns            boolean
    """
    if len(list(element)) == 0:
        return True
    else:
        return False


def derive_modifiers_and_string(item):
    """ Derive Modifiers and String

    At this point it is assumed that the parsed XML element is a UI string with some modifiers (<b></b>, <u></u>,
    etc). This function move through all XML layers while building a csv string that describes the XML modifiers
    that are found. At the deepest layer, the UI string is taken and both string are then returned

    :param item:                Parsed XML element.
    :return modifiers_string:   CSV string that describe UI string modifiers.
    :return ui_string:          The UI string.
    """
    number_of_layers = 0
    modifiers_string = ""
    ui_string = ""
    base_layer_found = False
    current_level = item

    while not base_layer_found:
        number_of_layers += 1
        if is_deepest_item(current_level):
            base_layer_found = True
        else:
            current_level = list(current_level)[0]

    current_level = item

    for index in range(number_of_layers):
        if index == 1:
            modifiers_string += current_level.tag
        elif index > 1:
            modifiers_string += ",{}".format(current_level.tag)

        if index == number_of_layers-1:
            ui_string = current_level.text
        else:
            current_level = list(current_level)[0]

    return modifiers_string, ui_string


def style_worksheet(worksheet):
    """ Style Worksheet

    Calls multiple functions to style the parsed worksheet.

    :param worksheet:   Parsed openPyXl worksheet.
    :returns worksheet: Worksheet has now been styled.
    """
    build_headings(worksheet)
    adjust_column_width(worksheet)
    apply_borders_and_colour(worksheet, "90CAF9", "BBDEFB")

    logger.info("Style elements successfully applied to Excel file.")


def build_headings(worksheet):
    """ Build Headings

    Prints hardcoded headings onto the openPyXl worksheet.

    :param worksheet:   openPyXl worksheet.
    :return:
    """
    worksheet['A1'] = "XML Element Type"
    worksheet['B1'] = "String Style Modifiers"
    worksheet['C1'] = "XML Element Key"
    worksheet['D1'] = "English"
    worksheet['E1'] = "French"
    worksheet['F1'] = "Spanish"

    logger.info("Headings successfully applied to Excel file.")


def adjust_column_width(worksheet):
    """ Adjust Column Width

    Iterates through each column and adjusts the width to ensure there is no text overlap in worksheet. All language
    are set to a fixed width and configured to wrap text.

    :param worksheet:   openPyXl worksheet.
    :return:
    """
    for index in range(3):
        col = worksheet[get_column_value(index)]
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
            except Exception:
                logger.warning("An error has occurred during styling spreadsheet, this may affect its appearance.")
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    wrap_alignment = Alignment(wrap_text=True)
    for index in range(3, 30):
        col = worksheet[get_column_value(index)]
        column = col[0].column
        worksheet.column_dimensions[column].width = 50
        for cell in col:
            cell.alignment = wrap_alignment

    logger.info("Excel column widths adjusted.")


def apply_borders_and_colour(worksheet, heading_colour, fill_colour):
    """ Apply Borders and Colour

    Applies borders and coluring to the parsed openPyXl worksheet.

    :param worksheet:       openPyXl worksheet.
    :param heading_colour:  String describing the hex code of the colour used to fill spreadsheet headings.
    :param fill_colour:     String describing the hex code of the colour used to fill spreadsheet body.
    :return:
    """

    height = worksheet.max_row
    width = worksheet.max_column

    for style_col in range(width):
        column = get_column_value(style_col)
        style_cell_name = "{}{}".format(column, 1)
        worksheet[style_cell_name].font = Font(size=11, bold=True, color='FF424242')

    for style_col in range(width):
        column = get_column_value(style_col)
        for style_row in range(height):
            style_cell_name = "{}{}".format(column, style_row + 1)
            style_cell = worksheet[style_cell_name]

            if style_row != 0:
                if style_row % 2 == 0:
                    style_cell.fill = PatternFill("solid", fgColor=fill_colour)
            else:
                style_cell.fill = PatternFill("solid", fgColor=heading_colour)

            if style_row == 0 and style_col == 0:
                style_cell.border = Border(left=Side(border_style="thick", color='FF000000'),
                                           right=Side(border_style="thin", color='FF000000'),
                                           top=Side(border_style="thick", color='FF000000'),
                                           bottom=Side(border_style="thick", color='FF000000'))
            elif style_row == 0 and style_col == width - 1:
                style_cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                           right=Side(border_style="thick", color='FF000000'),
                                           top=Side(border_style="thick", color='FF000000'),
                                           bottom=Side(border_style="thick", color='FF000000'))
            elif style_row == 0:
                style_cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                           right=Side(border_style="thin", color='FF000000'),
                                           top=Side(border_style="thick", color='FF000000'),
                                           bottom=Side(border_style="thick", color='FF000000'))
            elif style_row == height - 1 and style_col == 0:
                style_cell.border = Border(left=Side(border_style="thick", color='FF000000'),
                                           right=Side(border_style="thin", color='FF000000'),
                                           bottom=Side(border_style="thick", color='FF000000'))
            elif style_row == height - 1 and style_col == width - 1:
                style_cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                           right=Side(border_style="thick", color='FF000000'),
                                           bottom=Side(border_style="thick", color='FF000000'))
            elif style_row == height - 1:
                style_cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                           right=Side(border_style="thin", color='FF000000'),
                                           bottom=Side(border_style="thick", color='FF000000'))
            elif style_col == width - 1:
                style_cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                           right=Side(border_style="thick", color='FF000000'),
                                           bottom=Side(border_style="dashed", color='FF000000'))
            elif style_col == 0:
                style_cell.border = Border(left=Side(border_style="thick", color='FF000000'),
                                           right=Side(border_style="thin", color='FF000000'),
                                           bottom=Side(border_style="dashed", color='FF000000'))
            else:
                style_cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                           right=Side(border_style="thin", color='FF000000'),
                                           bottom=Side(border_style="dashed", color='FF000000'))
    logger.info("Borders and colour successfully applied to Excel file.")


def get_column_value(index):
    """
    Takes a numerical column index and returns the alphabetical equivalent index suitable for Excel.
    Determined by calculating the quotient and remainder of the index.

    Ex. Index 27 = AA

    :param index:   The column position of the next value to be written to an Excel cell
    :return:        The Excel column position of the Excel cell where the value should be written
    """
    try:
        alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        column_value = ""
        result = "not yet populated"

        result = int(math.modf(index / len(alphabet))[1])
        if result > 0:
            column_value += alphabet[result - 1]
        column_value += alphabet[index % len(alphabet)]

        return column_value
    except Exception as exp:
        print("ERROR: result = {}, index = {}".format(result, index))


def create_xml_file(worksheet, col_number, rows, xml_tree, path):
    """ Create XML File

    Using the parsed openPyXl worksheet and column number to create a strings.xml file in the language which that column
    represents. The XML file is then stored at the parsed file path.

    :param worksheet:   openPyXl worksheet.
    :param col_number:  The column number of the language to be created.
    :param rows:        The number of rows within the openPyXl worksheet
    :param xml_tree:    The XML object to place the information into.
    :param path:        The path where the output file is to be saved.
    :return:
    """
    current_type = "string"
    for row in range(2, rows):
        modifier_string = worksheet["B{}".format(row)].value

        # Update current element type ('item' falls under string-array or plural)
        if worksheet["A{}".format(row)].value != "item":
            current_type = worksheet["A{}".format(row)].value

        if current_type == "string":
            # String element with no modifiers
            if modifier_string is None:
                string_element = elementTree.Element('string')
                string_element.set("name", str(worksheet["C{}".format(row)].value))
                string_element.text = worksheet["{}{}".format(get_column_value(col_number), row)].value
            # String element with string modifiers
            else:
                temp_value = str(worksheet["{}{}".format(get_column_value(col_number), row)].value)
                string_element = create_modified_element("string", modifier_string.split(","),
                                                         str(worksheet["C{}".format(row)].value), temp_value)
            xml_tree.append(string_element)

        elif str(worksheet["A{}".format(row)].value) == "item":
            # UI string has no modifiers
            if modifier_string is None:
                item_element = elementTree.Element("item")
                item_element.text = worksheet["{}{}".format(get_column_value(col_number), row)].value
            # UI has modifiers and therefore need to nest UI string in modifier xml elements
            else:
                temp_value = str(worksheet["{}{}".format(get_column_value(col_number), row)].value)
                item_element = create_modified_element("item", modifier_string.split(","),
                                                       str(worksheet["C{}".format(row)].value), temp_value)
            multiple_item_element.append(item_element)

            # plural element, therefore need to add 'quantity' tag and value
            if current_type == "plurals":
                item_element.set("quantity", str(worksheet["C{}".format(row)].value))

            #
            if str(worksheet["A{}".format(row+1)].value) != "item" or row == rows-1:
                xml_tree.append(multiple_item_element)

        elif current_type == "string-array":
            multiple_item_element = elementTree.Element('string-array')
            multiple_item_element.set("name", str(worksheet["C{}".format(row)].value))

        elif current_type == "plurals":
            multiple_item_element = elementTree.Element('plurals')
            multiple_item_element.set("name", str(worksheet["C{}".format(row)].value))
        else:
            logger.warning("Found unknown XML type: \"{}\" Element has not been added.".
                           format(str(worksheet["{}1".format(get_column_value(row))].value)))

        xml_string = elementTree.tostring(xml_tree)

        xml = minidom.parseString(xml_string)
        formatted_xml_string = xml.toprettyxml()
        thing = 1
    save_xml_file(path, xml_tree)


def create_modified_element(element_type, modifiers, key, text):
    """ Create Modified Element

    Creates a XML element that contains string modifiers like <b></b>, <u></u>, etc.

    :param element_type:    The type of XML element that is being created.
    :param modifiers:       List of strings that describes modifiers
    :param key:             The key that should be used in element.
    :param text:            The UI value of element
    :return:
    """
    base_element = elementTree.Element(element_type)
    if element_type != "item":
        base_element.set("name", key)
    number_of_modifiers = len(modifiers)
    count = 1
    for modifier in modifiers:
        if number_of_modifiers == 1:
            modifier_element = elementTree.Element(modifier)
            modifier_element.text = text
        elif count == 1:
            modifier_element = elementTree.Element(modifier)
        elif count == number_of_modifiers:
            child_element_list = list(modifier_element.iter())
            last_element = elementTree.Element(modifier)
            last_element.text = text
            child_element_list.pop(len(child_element_list) - 1).append(last_element)
        else:
            child_element_list = list(modifier_element.iter())
            elementTree.SubElement(child_element_list.pop(len(child_element_list) - 1), modifier)
        count += 1
    base_element.append(modifier_element)

    return base_element


def save_xml_file(path, xml_tree):
    """ Save XML File

    Converts xml_tree into structured string and then saves as "strings.xml"

    :param path:        The path where file should be saved.
    :param xml_tree:    XML object containing information to be saved.
    :return:
    """
    try:
        xml_string = elementTree.tostring(xml_tree)

        xml = minidom.parseString(xml_string)
        formatted_xml_string = xml.toprettyxml()

        with open(os.path.join(path, "strings.xml"), 'w', encoding='utf-8') as file:
            file.write(formatted_xml_string)

        logger.info("strings.xml file successfully saved at: {}".format(path))

    except Exception as exception:
        error_string_one = str(repr(exception))
        error_string_two = str(exception.args)
        error_string_three = str(traceback.format_exc())
        logger.debug(error_string_one + "\n" + error_string_two + "\n" + error_string_three)
        logger.error("There was an error while saving strings.xml file at: {} ".format(path))
        exit(1)


if __name__ == '__main__':
    arguments = parser.parse_args()
    if len(sys.argv) >= 2:
        main(arguments)
    else:
        logger.warning("Incorrect number of arguments provided")
        parser.print_help()
